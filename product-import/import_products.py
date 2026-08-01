#!/usr/bin/env python3
"""Generate or update HLC product pages from the Excel product import workbook."""

from __future__ import annotations

import argparse
import html
import json
import re
import shutil
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from openpyxl import load_workbook


BASIC_SHEET = "产品基础信息"
DETAIL_SHEET = "产品细节"
CONTENT_SHEET = "产品内容与SEO"
IMAGE_SHEET = "产品图片"
RELATED_SHEET = "相关产品"
HEADER_ROW = 4
DATA_START_ROW = 5
SITE_ORIGIN = "https://hlctex.com"
DIRECTORY_HEADER = "所属目录"
BAMBOO_DIRECTORY = "竹纤维面料"
LIQUID_AMMONIA_DIRECTORY = "丝光&液氨"
FUNCTIONAL_DIRECTORY = "功能性面料"
WOOL_DIRECTORY = "羊毛面料"
WOMENSWEAR_DIRECTORY = "女装面料"
EMBROIDERED_DIRECTORY = "绣花面料"

PRODUCT_CATALOGS = {
    BAMBOO_DIRECTORY: {
        "path": Path("textile/bamboo-fabric/index.html"),
        "label": "Bamboo Fabrics",
        "collection": "HLC BAMBOO KNIT COLLECTION",
    },
    LIQUID_AMMONIA_DIRECTORY: {
        "path": Path("textile/mercerized-liquid-ammonia-fabric/index.html"),
        "label": "Mercerized & Liquid Ammonia Fabrics",
        "collection": "HLC MERCERIZED & LIQUID AMMONIA COLLECTION",
    },
    FUNCTIONAL_DIRECTORY: {
        "path": Path("textile/functional/index.html"),
        "label": "Functional Fabrics",
        "collection": "HLC FUNCTIONAL KNIT COLLECTION",
    },
    WOOL_DIRECTORY: {
        "path": Path("textile/wool-fabric/index.html"),
        "label": "Wool Fabrics",
        "collection": "HLC WOOL FABRIC COLLECTION",
    },
    WOMENSWEAR_DIRECTORY: {
        "path": Path("textile/womenswear-fabric/index.html"),
        "label": "Womenswear Fabrics",
        "collection": "HLC WOMENSWEAR FABRIC COLLECTION",
    },
    EMBROIDERED_DIRECTORY: {
        "path": Path("textile/embroidered-fabric/index.html"),
        "label": "Embroidered Fabrics",
        "collection": "HLC EMBROIDERED FABRIC COLLECTION",
    },
}

FIELDS = {
    "breadcrumb": "产品名称（英文）",
    "collection": "系列名称（英文）",
    "product-name": "产品名称（英文）",
    "style-number": "款号（Style#）",
    "yard-price": "实时价格 USD/码",
    "kg-price": "实时价格 USD/KG",
    "composition": "成分",
    "weight": "克重（g/m²）",
    "width": "有效幅宽（cm）",
    "construction": "织物组织",
}

CONTENT_FIELDS = {
    "short-description": "产品描述（英文）",
    "other-details": "其他信息（英文）",
    "seo-kicker": "SEO分类标签（英文）",
    "seo-heading": "SEO主标题（英文）",
    "seo-paragraph-1": "SEO段落1（英文）",
    "seo-paragraph-2": "SEO段落2（英文）",
}

TEST_RESULT_TEXT_HEADER = "测试结果（英文）"
TEST_RESULT_IMAGE_HEADER = "测试结果图片路径（可选）"
TEST_RESULT_ALT_HEADER = "测试结果图片ALT（英文）"
TEST_RESULT_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
TEST_RESULT_SOURCE_EXTENSIONS = TEST_RESULT_IMAGE_EXTENSIONS | {".pdf"}

ALT_FIELDS = [
    "主图ALT（英文）",
    "图2 ALT（英文）",
    "图3 ALT（英文）",
    "图4 ALT（英文）",
]

IMAGE_HEADERS = [
    "主图（必填）",
    "图2",
    "图3",
    "图4",
    "SEO内容图",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def first_nonempty(*values: Any) -> Any:
    for value in values:
        if clean(value):
            return value
    return ""


def slug_is_valid(slug: str) -> bool:
    return bool(re.fullmatch(r"[a-z0-9]+(?:-[a-z0-9]+)*", slug))


def product_name_from_slug(slug: str) -> str:
    name = slug.replace("-", " ").title()
    if name.startswith("Bamboo Spandex "):
        name = name.replace("Bamboo Spandex ", "Bamboo Viscose Spandex ", 1)
    return name


def read_rows(sheet) -> tuple[dict[str, dict[str, Any]], dict[int, str], dict[int, str]]:
    header_row = HEADER_ROW
    for candidate in range(1, min(sheet.max_row, 12) + 1):
        candidate_values = {
            clean(cell.value)
            for cell in sheet[candidate]
            if clean(cell.value)
        }
        if "产品URL标识" in candidate_values:
            header_row = candidate
            break
    headers = {
        cell.column: clean(cell.value)
        for cell in sheet[header_row]
        if clean(cell.value)
    }
    rows: dict[str, dict[str, Any]] = {}
    row_slug: dict[int, str] = {}
    col_header = dict(headers)
    for row_index in range(header_row + 1, sheet.max_row + 1):
        row: dict[str, Any] = {}
        for column, header in headers.items():
            cell = sheet.cell(row=row_index, column=column)
            value = cell.value
            if cell.data_type == "f" and isinstance(value, str):
                product_slug = clean(sheet[f"B{row_index}"].value)
                if header == "产品名称（英文）" and slug_is_valid(product_slug):
                    value = product_name_from_slug(product_slug)
                elif reference := re.fullmatch(
                    r"=\$?([A-Z]{1,3})\$?(\d+)", value.strip()
                ):
                    value = sheet[f"{reference.group(1)}{reference.group(2)}"].value
            row[header] = value
        slug = clean(row.get("产品URL标识"))
        if not slug:
            continue
        rows[slug] = row
        row_slug[row_index] = slug
    return rows, row_slug, col_header


def align_companion_rows(
    primary_row_slug: dict[int, str],
    rows: dict[str, dict[str, Any]],
    row_slug: dict[int, str],
) -> tuple[dict[str, dict[str, Any]], dict[int, str]]:
    """Match companion sheets to the basic-information row when their slug is stale."""
    aligned_rows = dict(rows)
    aligned_row_slug = dict(row_slug)
    for row_number, primary_slug in primary_row_slug.items():
        if primary_slug in aligned_rows:
            aligned_row_slug[row_number] = primary_slug
            continue
        companion_slug = aligned_row_slug.get(row_number)
        if companion_slug and companion_slug in aligned_rows:
            aligned_rows[primary_slug] = aligned_rows[companion_slug]
            aligned_row_slug[row_number] = primary_slug
    return aligned_rows, aligned_row_slug


def image_extension(image) -> str:
    fmt = clean(getattr(image, "format", "")).lower()
    if fmt == "jpeg":
        return ".jpg"
    if fmt in {"png", "gif", "webp", "bmp", "tiff"}:
        return f".{fmt}"
    return ".png"


def normalize_asset_name(slot: str) -> str:
    names = {
        "主图（必填）": "main",
        "图2": "view-2",
        "图3": "view-3",
        "图4": "view-4",
        "SEO内容图": "editorial",
    }
    return names[slot]


def strip_wrapping_quotes(value: Any) -> str:
    raw = clean(value)
    if len(raw) >= 2 and raw[0] == raw[-1] and raw[0] in {'"', "'"}:
        return raw[1:-1].strip()
    return raw


def is_test_result_image_reference(value: Any) -> bool:
    raw = strip_wrapping_quotes(value)
    if not raw:
        return False
    path = urlparse(raw).path if raw.startswith(("https://", "http://")) else raw
    return Path(path).suffix.lower() in TEST_RESULT_SOURCE_EXTENSIONS


def find_pdftoppm() -> Path:
    executable_name = "pdftoppm.exe" if sys.platform == "win32" else "pdftoppm"
    bundled = (
        Path(sys.executable).resolve().parent.parent
        / "native"
        / "poppler"
        / "Library"
        / "bin"
        / executable_name
    )
    if bundled.exists():
        return bundled
    discovered = shutil.which("pdftoppm")
    if discovered:
        return Path(discovered)
    raise FileNotFoundError(
        "PDF test result requires Poppler pdftoppm to create a web-viewable image"
    )


def render_pdf_test_result(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    output_prefix = destination.with_suffix("")
    subprocess.run(
        [
            str(find_pdftoppm()),
            "-png",
            "-r",
            "300",
            "-singlefile",
            str(source),
            str(output_prefix),
        ],
        check=True,
    )
    if not destination.exists():
        raise FileNotFoundError(
            f"PDF conversion did not create the expected image: {destination}"
        )


def copy_or_reference_image(
    value: Any,
    workbook_dir: Path,
    asset_dir: Path,
    slug: str,
    slot: str,
) -> str:
    raw = strip_wrapping_quotes(value)
    if not raw:
        return ""
    if raw.startswith(("https://", "http://", "/")):
        return raw

    source = Path(raw)
    if not source.is_absolute():
        source = (workbook_dir / source).resolve()
    if not source.exists():
        raise FileNotFoundError(f"{slug}: image file not found: {source}")

    asset_dir.mkdir(parents=True, exist_ok=True)
    extension = source.suffix.lower() or ".jpg"
    destination = asset_dir / f"{normalize_asset_name(slot)}{extension}"
    shutil.copy2(source, destination)
    return f"/assets/products/{slug}/{destination.name}"


def copy_or_reference_test_result(
    value: Any,
    workbook_dir: Path,
    repo: Path,
    slug: str,
    dry_run: bool,
) -> str:
    raw = strip_wrapping_quotes(value)
    if not raw:
        return ""
    if not is_test_result_image_reference(raw):
        raise ValueError(
            f"{slug}: test result must be PDF, JPG, JPEG, PNG, WEBP or GIF"
        )
    if raw.startswith(("https://", "http://", "/")):
        if Path(urlparse(raw).path).suffix.lower() == ".pdf":
            raise ValueError(
                f"{slug}: remote PDF test result cannot be converted; use a local file"
            )
        return raw

    source = Path(raw)
    if not source.is_absolute():
        source = (workbook_dir / source).resolve()
    if not source.exists():
        raise FileNotFoundError(f"{slug}: test result image not found: {source}")

    source_extension = source.suffix.lower()
    extension = ".png" if source_extension == ".pdf" else source_extension
    destination = (
        repo
        / "assets"
        / "products"
        / slug
        / f"test-result{extension}"
    )
    if not dry_run:
        if source_extension == ".pdf":
            render_pdf_test_result(source, destination)
        else:
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, destination)
    return f"/assets/products/{slug}/{destination.name}"


def extract_images(
    workbook,
    workbook_path: Path,
    repo: Path,
    rows: dict[str, dict[str, Any]],
    row_slug: dict[int, str],
    col_header: dict[int, str],
    dry_run: bool,
) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {slug: {} for slug in rows}
    sheet = workbook[IMAGE_SHEET]

    for slug, row in rows.items():
        asset_dir = repo / "assets" / "products" / slug
        for slot in IMAGE_HEADERS:
            if slot not in row:
                continue
            if dry_run:
                raw = clean(row.get(slot))
                if raw:
                    result[slug][slot] = raw
                continue
            result[slug][slot] = copy_or_reference_image(
                row.get(slot),
                workbook_path.parent,
                asset_dir,
                slug,
                slot,
            )

    for image in getattr(sheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        row_number = marker.row + 1
        column_number = marker.col + 1
        slug = row_slug.get(row_number)
        slot = col_header.get(column_number)
        if not slug or slot not in IMAGE_HEADERS:
            continue
        extension = image_extension(image)
        destination = repo / "assets" / "products" / slug / (
            f"{normalize_asset_name(slot)}{extension}"
        )
        result.setdefault(slug, {})[slot] = (
            f"/assets/products/{slug}/{destination.name}"
        )
        if not dry_run:
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(image._data())
    return result


def set_text(soup: BeautifulSoup, field: str, value: Any) -> None:
    node = soup.find(attrs={"data-template-field": field})
    if node is not None and clean(value):
        node.clear()
        node.append(clean(value))


def set_or_remove_detail(soup: BeautifulSoup, field: str, value: Any) -> None:
    node = soup.find(attrs={"data-template-field": field})
    if node is None:
        return
    if clean(value):
        node.clear()
        node.append(clean(value))
        return
    row = node.find_parent(attrs={"data-detail-row": True})
    if row is not None:
        row.decompose()


def remove_product_tab(soup: BeautifulSoup, name: str) -> None:
    button = soup.find(attrs={"data-product-tab": name})
    panel = soup.find(attrs={"data-product-panel": name})
    if button is not None:
        button.decompose()
    if panel is not None:
        panel.decompose()


def render_test_results(
    soup: BeautifulSoup,
    *,
    image_source: str,
    text: str,
    image_alt: str,
    product_name: str,
) -> None:
    panel = soup.find(attrs={"data-product-panel": "testing"})
    if panel is None:
        return
    if not image_source and not text:
        remove_product_tab(soup, "testing")
        return

    panel.clear()
    if image_source:
        figure = soup.new_tag("figure", attrs={"class": "catalog-test-result-media"})
        link = soup.new_tag(
            "a",
            href=image_source,
            target="_blank",
            rel="noopener",
            attrs={"aria-label": "Open high-resolution test result image"},
        )
        link.append(
            soup.new_tag(
                "img",
                src=image_source,
                alt=image_alt or f"{product_name} fabric test result",
                loading="lazy",
                decoding="async",
            )
        )
        figure.append(link)
        panel.append(figure)
    if text:
        paragraph = soup.new_tag("p")
        paragraph["data-template-field"] = "test-results"
        paragraph.string = text
        panel.append(paragraph)


def ensure_meta(soup: BeautifulSoup, *, name: str | None = None, prop: str | None = None):
    attrs = {"name": name} if name else {"property": prop}
    node = soup.head.find("meta", attrs=attrs)
    if node is None:
        node = soup.new_tag("meta")
        for key, value in attrs.items():
            node[key] = value
        soup.head.append(node)
    return node


def ensure_link(soup: BeautifulSoup, rel: str):
    node = soup.head.find("link", rel=rel)
    if node is None:
        node = soup.new_tag("link", rel=rel)
        soup.head.append(node)
    return node


def format_date(value: Any) -> tuple[str, str]:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        display = value.strftime("%b %d, %Y").replace(" 0", " ")
        return display, value.isoformat()
    raw = clean(value)
    if not raw:
        return "", ""
    try:
        parsed = datetime.fromisoformat(raw).date()
        return parsed.strftime("%b %d, %Y").replace(" 0", " "), parsed.isoformat()
    except ValueError:
        return raw, raw


def build_gallery(
    soup: BeautifulSoup,
    image_map: dict[str, str],
    content_row: dict[str, Any],
    product_name: str,
) -> list[str]:
    sources = [
        image_map.get(slot, "")
        for slot in IMAGE_HEADERS[:4]
        if image_map.get(slot, "")
    ]
    if not sources:
        return []

    alts = [
        clean(content_row.get(field)) or f"{product_name} product view {index + 1}"
        for index, field in enumerate(ALT_FIELDS)
    ]
    main = soup.find(attrs={"data-gallery-main": True})
    main["src"] = sources[0]
    main["data-full-image"] = sources[0]
    main["alt"] = alts[0]

    thumbnails = soup.select_one(".catalog-thumbnails")
    thumbnails.clear()
    for index, source in enumerate(sources):
        button = soup.new_tag(
            "button",
            attrs={
                "class": "catalog-thumbnail" + (" is-active" if index == 0 else ""),
                "type": "button",
                "role": "tab",
                "aria-selected": "true" if index == 0 else "false",
                "data-gallery-thumb": "",
                "data-image": source,
                "data-full-image": source,
                "data-alt": alts[index],
            },
        )
        thumb = soup.new_tag("img", src=source, alt="")
        button.append(thumb)
        thumbnails.append(button)

    dialog_image = soup.find(attrs={"data-gallery-dialog-image": True})
    if dialog_image is not None:
        dialog_image["src"] = sources[0]
        dialog_image["alt"] = f"{product_name} high-resolution view"
    return sources


def related_product_data(
    slug: str,
    basics: dict[str, dict[str, Any]],
    content: dict[str, dict[str, Any]],
    image_maps: dict[str, dict[str, str]],
) -> dict[str, str]:
    if slug in basics:
        row = basics[slug]
        return {
            "href": f"/textile/products/{slug}/",
            "name": clean(row.get("产品名称（英文）")) or slug,
            "style": f"Style#: {clean(row.get('款号（Style#）'))}",
            "image": image_maps.get(slug, {}).get("主图（必填）", "/assets/bamboo-knit-hero.jpg"),
            "alt": clean(content.get(slug, {}).get("主图ALT（英文）")) or clean(row.get("产品名称（英文）")),
        }
    if slug in {"bvf", "bvcf"}:
        return {
            "href": f"/textile/bamboo-fabric/{slug}/",
            "name": slug.upper(),
            "style": "View established product",
            "image": "/assets/bamboo-knit-hero.jpg" if slug == "bvf" else "/assets/bvcf-knit-hero.png",
            "alt": f"{slug.upper()} bamboo knit fabric",
        }
    return {
        "href": f"/textile/products/{slug}/",
        "name": slug.replace("-", " ").title(),
        "style": "View product",
        "image": "/assets/bamboo-knit-hero.jpg",
        "alt": slug.replace("-", " "),
    }


def build_related_cards(
    soup: BeautifulSoup,
    related_slugs: list[str],
    basics: dict[str, dict[str, Any]],
    content: dict[str, dict[str, Any]],
    image_maps: dict[str, dict[str, str]],
) -> None:
    if not related_slugs:
        return
    grid = soup.select_one(".catalog-related-grid")
    if grid is None:
        return
    grid.clear()
    for slug in related_slugs[:3]:
        data = related_product_data(slug, basics, content, image_maps)
        card = soup.new_tag("a", href=data["href"], attrs={"class": "catalog-related-card"})
        image_span = soup.new_tag("span", attrs={"class": "catalog-related-image"})
        image_span.append(soup.new_tag("img", src=data["image"], alt=data["alt"]))
        card.append(image_span)
        for class_name, text in [
            ("catalog-related-name", data["name"]),
            ("catalog-related-meta", data["style"]),
            ("catalog-related-price", "View product"),
        ]:
            span = soup.new_tag("span", attrs={"class": class_name})
            span.string = text
            card.append(span)
        grid.append(card)


def bamboo_catalog_filter_value(composition: str) -> str:
    normalized = composition.lower()
    if "organic cotton" in normalized:
        return "bamboo-organic-cotton"
    if "cotton" in normalized and any(
        term in normalized for term in ["spandex", "elastane"]
    ):
        return "bamboo-cotton-spandex"
    if "cotton" in normalized:
        return "bamboo-cotton"
    if "polyester" in normalized:
        return "bamboo-polyester"
    if any(term in normalized for term in ["spandex", "elastane"]):
        return "bamboo-spandex"
    return "bamboo-viscose"


def bamboo_catalog_applications(value: Any) -> str:
    normalized = clean(value).lower()
    applications: list[str] = []
    if any(term in normalized for term in ["baby", "infant", "footie", "zippy"]):
        applications.append("babywear")
    if any(term in normalized for term in ["sleep", "pajama", "lounge"]):
        applications.append("sleepwear")
    return " ".join(applications)


def catalog_composition_filter_value(composition: Any, directory_name: str) -> str:
    """Return one stable catalogue-filter token for the leading fibre story."""
    normalized = clean(composition).lower()
    if directory_name == BAMBOO_DIRECTORY:
        return bamboo_catalog_filter_value(composition)
    if directory_name == WOOL_DIRECTORY:
        return "wool-blend" if any(
            term in normalized
            for term in ["cotton", "viscose", "polyester", "nylon", "spandex", "elastane"]
        ) else "wool"
    if directory_name == EMBROIDERED_DIRECTORY:
        if "cotton" in normalized and not any(
            term in normalized for term in ["polyester", "nylon", "viscose"]
        ):
            return "cotton"
        if any(term in normalized for term in ["polyester", "nylon"]) and "cotton" not in normalized:
            return "synthetic"
        return "blended"
    for token, terms in [
        ("bamboo-viscose", ["bamboo"]),
        ("wool", ["wool", "merino"]),
        ("cotton", ["cotton", "supima", "giza"]),
        ("tencel-lyocell", ["tencel", "lyocell"]),
        ("viscose-modal", ["viscose", "modal"]),
        ("linen-ramie", ["linen", "flax", "ramie"]),
        ("recycled", ["recycled"]),
        ("polyester", ["polyester"]),
        ("nylon", ["nylon", "polyamide"]),
    ]:
        if any(term in normalized for term in terms):
            return token
    return "blended"


def catalog_applications(value: Any) -> str:
    normalized = clean(value).lower()
    groups = [
        ("babywear", ["baby", "infant", "zippy", "footie", "swaddle"]),
        ("sleepwear", ["sleep", "pajama", "lounge", "homewear"]),
        ("activewear", ["active", "sport", "legging", "running", "gym"]),
        ("womenswear", ["women", "dress", "skirt", "blouse", "fashion"]),
        ("menswear", ["men", "polo", "shirt"]),
        ("kidswear", ["kid", "child"]),
        ("underwear", ["underwear", "lingerie", "base layer"]),
        ("outerwear", ["outerwear", "jacket", "coat"]),
    ]
    return " ".join(
        token
        for token, terms in groups
        if any(term in normalized for term in terms)
    )


def refresh_catalog_filters(soup: BeautifulSoup, cards: list[Any]) -> None:
    """Rebuild filter choices from the products currently published in a catalogue."""
    for details in soup.select(".bamboo-catalog-filters details"):
        input_element = details.select_one("[data-filter]")
        fieldset = details.find("fieldset")
        if input_element is None or fieldset is None:
            continue
        filter_name = clean(input_element.get("data-filter"))
        values: dict[str, int] = {}
        for card in cards:
            for value in clean(card.get(f"data-{filter_name}")).split():
                if value:
                    values[value] = values.get(value, 0) + 1
        legend = fieldset.find("legend")
        legend_text = legend.get_text(" ", strip=True) if legend is not None else f"{filter_name.title()} options"
        fieldset.clear()
        new_legend = soup.new_tag("legend")
        new_legend.string = legend_text
        fieldset.append(new_legend)
        for value, count in sorted(values.items()):
            label = soup.new_tag("label")
            checkbox = soup.new_tag(
                "input",
                attrs={
                    "type": "checkbox",
                    "value": value,
                    "data-filter": filter_name,
                },
            )
            label.append(checkbox)
            label.append(
                " "
                + (
                    f"{value} g/m²"
                    if filter_name == "weight"
                    else value.replace("-", " ").title()
                )
                + " "
            )
            count_element = soup.new_tag("span")
            count_element.string = str(count)
            label.append(count_element)
            fieldset.append(label)


def update_bamboo_catalog(
    *,
    repo: Path,
    basics: dict[str, dict[str, Any]],
    contents: dict[str, dict[str, Any]],
    image_maps: dict[str, dict[str, str]],
    dry_run: bool,
) -> None:
    """Synchronize workbook products assigned to the bamboo fabric directory."""
    catalog_path = repo / "textile" / "bamboo-fabric" / "index.html"
    soup = BeautifulSoup(catalog_path.read_text(encoding="utf-8"), "html.parser")
    grid = soup.select_one("[data-product-grid]")
    if grid is None:
        raise ValueError("Bamboo fabric catalog product grid not found")
    if not any(DIRECTORY_HEADER in row for row in basics.values()):
        return

    workbook_slugs = set(basics)
    existing_cards: dict[str, Any] = {}
    for card in list(grid.select("[data-product-card]")):
        link = card.select_one('a[href^="/textile/products/"]')
        if link is None:
            continue
        match = re.fullmatch(r"/textile/products/([^/]+)/", clean(link.get("href")))
        if not match:
            continue
        slug = match.group(1)
        if slug in workbook_slugs:
            existing_cards[slug] = card

    if not any(
        clean(row.get("发布")).upper() == "YES"
        and (
            clean(row.get(DIRECTORY_HEADER)) == BAMBOO_DIRECTORY
            or slug in existing_cards
        )
        for slug, row in basics.items()
    ):
        return

    featured = len(grid.select("[data-product-card]"))
    for slug, basic in basics.items():
        existing_card = existing_cards.get(slug)
        is_publishing = clean(basic.get("发布")).upper() == "YES"
        if not is_publishing:
            continue
        should_list = clean(basic.get(DIRECTORY_HEADER)) == BAMBOO_DIRECTORY
        if not should_list:
            if existing_card is not None:
                existing_card.decompose()
            continue

        product_name = clean(basic.get("产品名称（英文）")) or slug.replace("-", " ").title()
        composition = clean(basic.get("成分"))
        weight = clean(basic.get("克重（g/m²）"))
        construction = clean(basic.get("织物组织"))
        applications = bamboo_catalog_applications(basic.get("适用产品"))
        image = image_maps.get(slug, {}).get("主图（必填）")
        if not image:
            continue
        content = contents.get(slug, {})
        alt = clean(content.get("主图ALT（英文）")) or f"{product_name} fabric"
        card_featured = (
            clean(existing_card.get("data-featured"))
            if existing_card is not None
            else ""
        )
        if not card_featured:
            featured += 1
            card_featured = str(featured)

        card = soup.new_tag(
            "article",
            attrs={
                "class": "bamboo-product-card",
                "data-product-card": "",
                "data-featured": card_featured,
                "data-name": product_name,
                "data-price": clean(basic.get("实时价格 USD/码")),
                "data-composition": bamboo_catalog_filter_value(composition),
                "data-weight": weight,
                "data-construction": re.sub(
                    r"[^a-z0-9]+",
                    "-",
                    construction.lower(),
                ).strip("-"),
                "data-application": applications,
            },
        )
        href = f"/textile/products/{slug}/"
        image_link = soup.new_tag("a", href=href, attrs={"class": "bamboo-product-image"})
        image_tag = soup.new_tag(
            "img",
            src=image,
            alt=alt,
            width="1280",
            height="1280",
            loading="lazy",
            decoding="async",
        )
        image_link.append(image_tag)
        overlay = soup.new_tag("span")
        overlay.string = "View fabric"
        image_link.append(overlay)
        card.append(image_link)

        info = soup.new_tag("div", attrs={"class": "bamboo-product-info"})
        family = soup.new_tag("p", attrs={"class": "bamboo-product-family"})
        family.string = clean(basic.get("系列名称（英文）")) or "HLC BAMBOO KNIT COLLECTION"
        info.append(family)
        heading = soup.new_tag("h2")
        heading_link = soup.new_tag("a", href=href)
        heading_link.string = product_name
        heading.append(heading_link)
        info.append(heading)
        style = soup.new_tag("p", attrs={"class": "bamboo-product-style"})
        style.string = f"Style#: {clean(basic.get('款号（Style#）'))}"
        info.append(style)
        spec = soup.new_tag("p", attrs={"class": "bamboo-product-spec"})
        spec.string = f"{composition} · {weight} g/m²"
        info.append(spec)
        price = soup.new_tag("p", attrs={"class": "bamboo-product-price"})
        price_yard = soup.new_tag("strong")
        price_yard.string = f"US${clean(basic.get('实时价格 USD/码'))}"
        price.append(price_yard)
        price.append("/yd ")
        price_kg = soup.new_tag("span")
        price_kg.string = f"US${clean(basic.get('实时价格 USD/KG'))}/kg"
        price.append(price_kg)
        info.append(price)
        card.append(info)
        if existing_card is not None:
            existing_card.replace_with(card)
        else:
            grid.append(card)

    cards = grid.select("[data-product-card]")
    result_count = soup.select_one("[data-result-count]")
    if result_count is not None:
        result_count.string = str(len(cards))

    for input_element in soup.select("[data-filter]"):
        filter_name = clean(input_element.get("data-filter"))
        filter_value = clean(input_element.get("value"))
        count = sum(
            filter_value
            in clean(card.get(f"data-{filter_name}")).split()
            for card in cards
        )
        label = input_element.find_parent("label")
        count_element = label.find("span") if label is not None else None
        if count_element is not None:
            count_element.string = str(count)

    for schema_tag in soup.find_all("script", type="application/ld+json"):
        try:
            structured_data = json.loads(schema_tag.string or "")
        except json.JSONDecodeError:
            continue
        graph = structured_data.get("@graph")
        if not isinstance(graph, list):
            continue
        item_list = next(
            (
                node
                for node in graph
                if isinstance(node, dict) and node.get("@type") == "ItemList"
            ),
            None,
        )
        if item_list is None:
            continue
        item_list["numberOfItems"] = len(cards)
        item_list["itemListElement"] = [
            {
                "@type": "ListItem",
                "position": position,
                "url": SITE_ORIGIN + clean(card.select_one("h2 a").get("href")),
                "name": card.select_one("h2 a").get_text(" ", strip=True),
            }
            for position, card in enumerate(cards, start=1)
            if card.select_one("h2 a") is not None
        ]
        schema_tag.string = json.dumps(structured_data, ensure_ascii=False, indent=2)
        break

    if not dry_run:
        catalog_path.write_text("<!doctype html>\n" + str(soup.html), encoding="utf-8")


def update_liquid_ammonia_catalog(
    *,
    repo: Path,
    basics: dict[str, dict[str, Any]],
    contents: dict[str, dict[str, Any]],
    image_maps: dict[str, dict[str, str]],
    dry_run: bool,
) -> None:
    """Add or update YES products without removing existing or NO product cards."""
    assigned = [
        (slug, row)
        for slug, row in basics.items()
        if clean(row.get("发布")).upper() == "YES"
        and clean(row.get(DIRECTORY_HEADER)) == LIQUID_AMMONIA_DIRECTORY
    ]
    if not assigned:
        return

    catalog_path = repo / "pickup" / "mercerization-liquid-ammonia" / "index.html"
    updated = catalog_path.read_text(encoding="utf-8")
    grid_marker = '<div class="business-grid related-product-grid">'
    grid_start = updated.find(grid_marker)
    if grid_start < 0:
        raise ValueError("Liquid-ammonia directory product grid not found")

    for slug, basic in assigned:
        image = image_maps.get(slug, {}).get("主图（必填）")
        if not image:
            continue
        content = contents.get(slug, {})
        product_name = clean(basic.get("产品名称（英文）")) or slug.replace("-", " ").title()
        alt = clean(content.get("主图ALT（英文）")) or f"{product_name} fabric"
        style = clean(basic.get("款号（Style#）"))
        price = clean(basic.get("实时价格 USD/码"))
        card = (
            f'<a class="related-product-card" href="/textile/products/{slug}/" '
            f'aria-label="View {html.escape(product_name)} details">'
            f'<img src="{html.escape(image)}" width="1000" height="750" loading="lazy" '
            f'decoding="async" alt="{html.escape(alt)}">'
            '<span class="related-product-overlay"><small>LIQUID AMMONIA</small>'
            f'<h3>{html.escape(product_name)}</h3>'
            '<span class="related-product-link">View details '
            '<span aria-hidden="true">→</span></span></span>'
            '<span class="related-product-meta">'
            f'<span>Style No.<strong>{html.escape(style)}</strong></span>'
            f'<span>Price<strong>US${html.escape(price)}/yd</strong></span>'
            '</span></a>'
        )
        existing_pattern = re.compile(
            rf'<a class="related-product-card" '
            rf'href="/textile/products/{re.escape(slug)}/"[^>]*>.*?</a>',
            flags=re.DOTALL,
        )
        if existing_pattern.search(updated, grid_start):
            updated = existing_pattern.sub(lambda _: card, updated, count=1)
            continue
        grid_end = updated.find("\n        </div>", grid_start)
        if grid_end < 0:
            raise ValueError("Liquid-ammonia directory product grid closing tag not found")
        updated = updated[:grid_end] + f"\n          {card}" + updated[grid_end:]

    if not dry_run:
        catalog_path.write_text(updated, encoding="utf-8")


def update_product_catalog(
    *,
    repo: Path,
    basics: dict[str, dict[str, Any]],
    contents: dict[str, dict[str, Any]],
    image_maps: dict[str, dict[str, str]],
    directory_name: str,
    catalog_relative_path: Path,
    default_collection: str,
    dry_run: bool,
) -> None:
    """Synchronize one workbook directory with its independent product catalog."""
    catalog_path = repo / catalog_relative_path
    soup = BeautifulSoup(catalog_path.read_text(encoding="utf-8"), "html.parser")
    grid = soup.select_one("[data-product-grid]")
    if grid is None:
        raise ValueError(f"{directory_name} product catalog grid not found")
    if not any(DIRECTORY_HEADER in row for row in basics.values()):
        return

    workbook_slugs = set(basics)
    existing_cards: dict[str, Any] = {}
    for card in list(grid.select("[data-product-card]")):
        link = card.select_one('a[href^="/textile/products/"]')
        if link is None:
            continue
        match = re.fullmatch(r"/textile/products/([^/]+)/", clean(link.get("href")))
        if match and match.group(1) in workbook_slugs:
            existing_cards[match.group(1)] = card

    publish_header = next(
        (
            header
            for row in basics.values()
            for header in row
            if clean(row.get(header)).upper() in {"YES", "NO"}
        ),
        "",
    )
    if not publish_header:
        return
    if not any(
        clean(row.get(publish_header)).upper() == "YES"
        and (
            clean(row.get(DIRECTORY_HEADER)) == directory_name
            or slug in existing_cards
        )
        for slug, row in basics.items()
    ):
        return

    application_header = next(
        (
            header
            for row in basics.values()
            for header in row
            if any(
                token in clean(header).lower()
                for token in ["application", "适用", "閫傜敤"]
            )
        ),
        "",
    )
    featured = len(grid.select("[data-product-card]"))
    for slug, basic in basics.items():
        existing_card = existing_cards.get(slug)
        if clean(basic.get(publish_header)).upper() != "YES":
            continue
        should_list = clean(basic.get(DIRECTORY_HEADER)) == directory_name
        if not should_list:
            if existing_card is not None:
                existing_card.decompose()
            continue

        product_name = clean(basic.get(FIELDS["product-name"])) or slug.replace(
            "-", " "
        ).title()
        composition = clean(basic.get(FIELDS["composition"]))
        weight = clean(basic.get(FIELDS["weight"]))
        construction = clean(basic.get(FIELDS["construction"]))
        application_text = clean(basic.get(application_header))
        applications = catalog_applications(application_text)

        image = image_maps.get(slug, {}).get(IMAGE_HEADERS[0])
        if not image:
            continue
        content = contents.get(slug, {})
        alt = clean(content.get(ALT_FIELDS[0])) or f"{product_name} fabric"
        card_featured = (
            clean(existing_card.get("data-featured"))
            if existing_card is not None
            else ""
        )
        if not card_featured:
            featured += 1
            card_featured = str(featured)

        card = soup.new_tag(
            "article",
            attrs={
                "class": "bamboo-product-card",
                "data-product-card": "",
                "data-featured": card_featured,
                "data-name": product_name,
                "data-price": clean(basic.get(FIELDS["yard-price"])),
                "data-composition": catalog_composition_filter_value(
                    composition,
                    directory_name,
                ),
                "data-weight": weight,
                "data-construction": re.sub(
                    r"[^a-z0-9]+", "-", construction.lower()
                ).strip("-"),
                "data-application": applications,
            },
        )
        href = f"/textile/products/{slug}/"
        image_link = soup.new_tag(
            "a", href=href, attrs={"class": "bamboo-product-image"}
        )
        image_parent = str(Path(image).parent).replace("\\", "/")
        display_700 = f"{image_parent}/display-700.jpg"
        display_1000 = f"{image_parent}/display-1000.jpg"
        has_display_700 = (repo / display_700.lstrip("/")).exists()
        has_display_1000 = (repo / display_1000.lstrip("/")).exists()
        image_tag = soup.new_tag(
            "img",
            src=display_700 if has_display_700 else image,
            alt=alt,
            width="1000",
            height="750",
            loading="lazy",
            decoding="async",
        )
        if has_display_700 and has_display_1000:
            image_tag["srcset"] = f"{display_700} 700w, {display_1000} 1000w"
            image_tag["sizes"] = (
                "(max-width: 760px) calc(100vw - 45px), "
                "(max-width: 980px) calc((100vw - 292px) / 2), 30vw"
            )
        image_link.append(image_tag)
        overlay = soup.new_tag("span")
        overlay.string = "View fabric"
        image_link.append(overlay)
        card.append(image_link)

        info = soup.new_tag("div", attrs={"class": "bamboo-product-info"})
        family = soup.new_tag("p", attrs={"class": "bamboo-product-family"})
        family.string = (
            clean(basic.get(FIELDS["collection"]))
            or default_collection
        )
        info.append(family)
        heading = soup.new_tag("h2")
        heading_link = soup.new_tag("a", href=href)
        heading_link.string = product_name
        heading.append(heading_link)
        info.append(heading)
        style = soup.new_tag("p", attrs={"class": "bamboo-product-style"})
        style.string = f"Style#: {clean(basic.get(FIELDS['style-number']))}"
        info.append(style)
        spec = soup.new_tag("p", attrs={"class": "bamboo-product-spec"})
        spec.string = f"{composition} · {weight} g/m²"
        info.append(spec)
        price = soup.new_tag("p", attrs={"class": "bamboo-product-price"})
        price_yard = soup.new_tag("strong")
        price_yard.string = f"US${clean(basic.get(FIELDS['yard-price']))}"
        price.append(price_yard)
        price.append("/yd ")
        price_kg = soup.new_tag("span")
        price_kg.string = f"US${clean(basic.get(FIELDS['kg-price']))}/kg"
        price.append(price_kg)
        info.append(price)
        price_date_display, _ = format_date(basic.get("价格有效期"))
        if price_date_display:
            validity = soup.new_tag(
                "small", attrs={"class": "bamboo-product-price-validity"}
            )
            validity.string = f"Valid through {price_date_display}"
            info.append(validity)
        card.append(info)
        if existing_card is not None:
            existing_card.replace_with(card)
        else:
            grid.append(card)

    cards = grid.select("[data-product-card]")
    result_count = soup.select_one("[data-result-count]")
    if result_count is not None:
        result_count.string = str(len(cards))
    refresh_catalog_filters(soup, cards)
    for input_element in soup.select("[data-filter]"):
        filter_name = clean(input_element.get("data-filter"))
        filter_value = clean(input_element.get("value"))
        count = sum(
            filter_value in clean(card.get(f"data-{filter_name}")).split()
            for card in cards
        )
        label = input_element.find_parent("label")
        count_element = label.find("span") if label is not None else None
        if count_element is not None:
            count_element.string = str(count)

    for schema_tag in soup.find_all("script", type="application/ld+json"):
        try:
            structured_data = json.loads(schema_tag.string or "")
        except json.JSONDecodeError:
            continue
        graph = structured_data.get("@graph")
        if not isinstance(graph, list):
            continue
        item_list = next(
            (
                node
                for node in graph
                if isinstance(node, dict) and node.get("@type") == "ItemList"
            ),
            None,
        )
        if item_list is None:
            continue
        item_list["numberOfItems"] = len(cards)
        item_list["itemListElement"] = [
            {
                "@type": "ListItem",
                "position": position,
                "url": SITE_ORIGIN + clean(card.select_one("h2 a").get("href")),
                "name": card.select_one("h2 a").get_text(" ", strip=True),
            }
            for position, card in enumerate(cards, start=1)
            if card.select_one("h2 a") is not None
        ]
        schema_tag.string = json.dumps(
            structured_data, ensure_ascii=False, indent=2
        )
        break

    if not dry_run:
        catalog_path.write_text(
            "<!doctype html>\n" + str(soup.html), encoding="utf-8"
        )


def update_sitemap(sitemap_path: Path, url: str, last_modified: str, dry_run: bool) -> None:
    text = sitemap_path.read_text(encoding="utf-8")
    escaped_url = re.escape(url)
    block_pattern = re.compile(
        rf"(<url><loc>{escaped_url}</loc>.*?</url>)",
        flags=re.DOTALL,
    )
    match = block_pattern.search(text)
    if match:
        updated_block = re.sub(
            r"<lastmod>[^<]*</lastmod>",
            f"<lastmod>{last_modified}</lastmod>",
            match.group(1),
            count=1,
        )
        updated = text[: match.start()] + updated_block + text[match.end() :]
    else:
        entry = (
            f"  <url><loc>{html.escape(url)}</loc><lastmod>{last_modified}</lastmod>"
            "<changefreq>weekly</changefreq><priority>0.88</priority></url>\n"
        )
        updated = text.replace("</urlset>", entry + "</urlset>")
    if not dry_run:
        sitemap_path.write_text(updated, encoding="utf-8")


def generate_page(
    *,
    repo: Path,
    slug: str,
    basic: dict[str, Any],
    detail_row: dict[str, Any],
    content_row: dict[str, Any],
    image_map: dict[str, str],
    test_result_image: str,
    related_slugs: list[str],
    basics: dict[str, dict[str, Any]],
    contents: dict[str, dict[str, Any]],
    image_maps: dict[str, dict[str, str]],
    dry_run: bool,
) -> Path:
    template_path = repo / "textile" / "product-template" / "index.html"
    output_path = repo / "textile" / "products" / slug / "index.html"
    product_name = clean(basic.get("产品名称（英文）"))
    style_number = clean(basic.get("款号（Style#）"))
    canonical = f"{SITE_ORIGIN}/textile/products/{slug}/"

    soup = BeautifulSoup(template_path.read_text(encoding="utf-8"), "html.parser")
    catalog = PRODUCT_CATALOGS.get(clean(basic.get(DIRECTORY_HEADER)))
    if catalog:
        directory_link = soup.select_one(".catalog-breadcrumbs li:nth-of-type(3) a")
        if directory_link is not None:
            directory_link["href"] = "/" + str(catalog["path"].parent).replace("\\", "/") + "/"
            directory_link.string = str(catalog["label"])
    for field, header in FIELDS.items():
        value = basic.get(header)
        if field == "weight" and clean(value):
            value = f"{clean(value)} g/m²"
        elif field == "width" and clean(value):
            value = f"{clean(value)} cm"
        set_text(soup, field, value)
    for field, header in CONTENT_FIELDS.items():
        set_text(soup, field, content_row.get(header))

    raw_test_result = clean(content_row.get(TEST_RESULT_TEXT_HEADER))
    test_result_text = (
        ""
        if is_test_result_image_reference(raw_test_result)
        else raw_test_result
    )
    render_test_results(
        soup,
        image_source=test_result_image,
        text=test_result_text,
        image_alt=clean(content_row.get(TEST_RESULT_ALT_HEADER))
        or f"{product_name} laboratory test report",
        product_name=product_name,
    )

    detail_intro = clean(
        first_nonempty(
            detail_row.get("细节说明（英文，选填）"),
            content_row.get("详细信息（英文）"),
        )
    )
    if detail_intro:
        set_text(soup, "technical-details", detail_intro)
    else:
        details_intro = soup.find(attrs={"data-template-field": "technical-details"})
        if details_intro is not None:
            details_intro.decompose()

    yards_per_kg = clean(
        first_nonempty(
            detail_row.get("每KG等于多少码"),
            basic.get("每KG等于多少码"),
        )
    )
    detail_values = {
        "detail-yarn-count": first_nonempty(
            detail_row.get("纱支"),
            basic.get("纱支"),
        ),
        "detail-moq": first_nonempty(
            detail_row.get("MOQ / MCQ（英文）"),
            basic.get("起订量 / 单色起订量"),
        ),
        "detail-weight-conversion": f"1 KG = {yards_per_kg} YDS" if yards_per_kg else "",
        "detail-sample-lead": first_nonempty(
            detail_row.get("Sample lead time（英文）"),
            basic.get("样品交期（英文）"),
        ),
        "detail-bulk-lead": first_nonempty(
            detail_row.get("Bulk lead time（英文）"),
            basic.get("大货交期（英文）"),
        ),
        "detail-applications": first_nonempty(
            detail_row.get("Applications（英文）"),
            basic.get("适用产品"),
        ),
        "detail-finishing": first_nonempty(
            detail_row.get("后整理（英文）"),
            basic.get("后整理"),
        ),
    }
    for field, value in detail_values.items():
        set_or_remove_detail(soup, field, value)

    if not clean(content_row.get("其他信息（英文）")):
        remove_product_tab(soup, "other")

    set_text(soup, "currency", "US$")
    set_text(soup, "currency-kg", "US$")
    price_date_display, price_date_iso = format_date(basic.get("价格有效期"))
    price_label = soup.select_one(".catalog-price-label")
    if price_label is not None and price_date_display:
        price_label.clear()
        price_label.append("Current price ")
        emphasis = soup.new_tag("em")
        emphasis.string = f"(valid through {price_date_display})"
        price_label.append(emphasis)

    sources = build_gallery(soup, image_map, content_row, product_name)
    editorial_source = image_map.get("SEO内容图") or (sources[0] if sources else "")
    editorial = soup.find(attrs={"data-template-field": "editorial-image"})
    if editorial is not None and editorial_source:
        editorial["src"] = editorial_source
        editorial["alt"] = (
            clean(content_row.get("SEO内容图ALT（英文）"))
            or f"{product_name} fabric development detail"
        )

    seo_title = clean(content_row.get("SEO标题（英文）")) or f"{product_name} | HLC"
    meta_description = clean(content_row.get("Meta描述（英文）")) or clean(
        content_row.get("产品描述（英文）")
    )
    soup.title.string = seo_title
    ensure_meta(soup, name="description")["content"] = meta_description
    ensure_meta(soup, name="robots")["content"] = "index,follow,max-image-preview:large"
    ensure_link(soup, "canonical")["href"] = canonical
    ensure_meta(soup, prop="og:type")["content"] = "product"
    ensure_meta(soup, prop="og:title")["content"] = seo_title
    ensure_meta(soup, prop="og:description")["content"] = meta_description
    ensure_meta(soup, prop="og:url")["content"] = canonical
    if sources:
        ensure_meta(soup, prop="og:image")["content"] = SITE_ORIGIN + sources[0]
    ensure_meta(soup, name="twitter:card")["content"] = "summary_large_image"
    ensure_meta(soup, name="twitter:title")["content"] = seo_title
    ensure_meta(soup, name="twitter:description")["content"] = meta_description

    build_related_cards(
        soup,
        related_slugs,
        basics,
        contents,
        image_maps,
    )

    structured_data: dict[str, Any] = {
        "@context": "https://schema.org",
        "@type": "Product",
        "name": product_name,
        "sku": style_number,
        "description": meta_description,
        "url": canonical,
        "brand": {"@type": "Brand", "name": "HLC"},
        "manufacturer": {"@type": "Organization", "name": "HLC Group Co., Ltd."},
        "countryOfOrigin": clean(basic.get("原产国")) or "China",
        "material": clean(basic.get("成分")),
    }
    if sources:
        structured_data["image"] = [
            source if source.startswith("http") else SITE_ORIGIN + source
            for source in sources
        ]

    yard_price = basic.get("实时价格 USD/码")
    kg_price = basic.get("实时价格 USD/KG")
    if yard_price not in (None, "") or kg_price not in (None, ""):
        price_specifications = []
        if yard_price not in (None, ""):
            price_specifications.append({
                "@type": "UnitPriceSpecification",
                "price": float(yard_price),
                "priceCurrency": "USD",
                "unitCode": "YDQ",
            })
        if kg_price not in (None, ""):
            price_specifications.append({
                "@type": "UnitPriceSpecification",
                "price": float(kg_price),
                "priceCurrency": "USD",
                "unitCode": "KGM",
            })
        offer: dict[str, Any] = {
            "@type": "Offer",
            "url": canonical,
            "priceCurrency": "USD",
            "availability": "https://schema.org/InStock",
            "seller": {"@type": "Organization", "name": "HLC Group Co., Ltd."},
            "priceSpecification": price_specifications,
        }
        if yard_price not in (None, ""):
            offer["price"] = float(yard_price)
        else:
            offer["price"] = float(kg_price)
        if price_date_iso:
            offer["priceValidUntil"] = price_date_iso
        structured_data["offers"] = offer

    schema = soup.new_tag("script", type="application/ld+json")
    schema.string = json.dumps(structured_data, ensure_ascii=False, indent=2)
    soup.head.append(schema)

    if not dry_run:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("<!doctype html>\n" + str(soup.html), encoding="utf-8")
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate HLC product pages from an Excel workbook."
    )
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--repo", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only", help="Generate only one product URL slug")
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    repo = args.repo.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if not (repo / "textile" / "product-template" / "index.html").exists():
        raise FileNotFoundError("HLC product template not found in repository")

    workbook = load_workbook(workbook_path, data_only=False)
    for required_sheet in [
        BASIC_SHEET,
        CONTENT_SHEET,
        IMAGE_SHEET,
        RELATED_SHEET,
    ]:
        if required_sheet not in workbook.sheetnames:
            raise ValueError(f"Missing worksheet: {required_sheet}")

    basics, basic_row_slug, _ = read_rows(workbook[BASIC_SHEET])
    if DETAIL_SHEET in workbook.sheetnames:
        details, detail_row_slug, _ = read_rows(workbook[DETAIL_SHEET])
        details, _ = align_companion_rows(
            basic_row_slug,
            details,
            detail_row_slug,
        )
    else:
        details = {}
    contents, content_row_slug, _ = read_rows(workbook[CONTENT_SHEET])
    image_rows, image_row_slug, image_headers = read_rows(workbook[IMAGE_SHEET])
    related_rows, related_row_slug, _ = read_rows(workbook[RELATED_SHEET])
    contents, _ = align_companion_rows(
        basic_row_slug,
        contents,
        content_row_slug,
    )
    image_rows, image_row_slug = align_companion_rows(
        basic_row_slug,
        image_rows,
        image_row_slug,
    )
    related_rows, _ = align_companion_rows(
        basic_row_slug,
        related_rows,
        related_row_slug,
    )
    publish_slugs = {
        slug
        for slug, row in basics.items()
        if clean(row.get("发布")).upper() == "YES"
        and (not args.only or slug == args.only)
    }
    image_rows = {
        slug: row
        for slug, row in image_rows.items()
        if slug in publish_slugs
    }
    image_row_slug = {
        row_number: slug
        for row_number, slug in image_row_slug.items()
        if slug in publish_slugs
    }
    image_maps = extract_images(
        workbook,
        workbook_path,
        repo,
        image_rows,
        image_row_slug,
        image_headers,
        args.dry_run,
    )

    selected: list[str] = []
    errors: list[str] = []
    for slug, basic in basics.items():
        if clean(basic.get("发布")).upper() != "YES":
            continue
        if args.only and slug != args.only:
            continue
        if not slug_is_valid(slug):
            errors.append(f"{slug}: URL identifier must use lowercase letters, numbers and hyphens")
            continue
        content_row = contents.get(slug, {})
        required = {
            "Style#": basic.get("款号（Style#）"),
            "Product name": basic.get("产品名称（英文）"),
            "Composition": basic.get("成分"),
            "Product description": content_row.get("产品描述（英文）"),
            "SEO Title": content_row.get("SEO标题（英文）"),
            "Meta Description": content_row.get("Meta描述（英文）"),
        }
        missing = [label for label, value in required.items() if not clean(value)]
        if not image_maps.get(slug, {}).get("主图（必填）"):
            missing.append("Main image")
        if missing:
            errors.append(f"{slug}: missing {', '.join(missing)}")
            continue

        related_row = related_rows.get(slug, {})
        related_slugs = [
            candidate
            for header in [
                "相关产品1 URL标识",
                "相关产品2 URL标识",
                "相关产品3 URL标识",
            ]
            if (candidate := clean(related_row.get(header)))
            and (candidate in basics or candidate in {"bvf", "bvcf"})
        ]
        output = generate_page(
            repo=repo,
            slug=slug,
            basic=basic,
            detail_row=details.get(slug, {}),
            content_row=content_row,
            image_map=image_maps.get(slug, {}),
            test_result_image=copy_or_reference_test_result(
                first_nonempty(
                    content_row.get(TEST_RESULT_IMAGE_HEADER),
                    content_row.get(TEST_RESULT_TEXT_HEADER)
                    if is_test_result_image_reference(
                        content_row.get(TEST_RESULT_TEXT_HEADER)
                    )
                    else "",
                ),
                workbook_path.parent,
                repo,
                slug,
                args.dry_run,
            ),
            related_slugs=related_slugs,
            basics=basics,
            contents=contents,
            image_maps=image_maps,
            dry_run=args.dry_run,
        )
        selected.append(slug)
        print(f"{'[DRY RUN] ' if args.dry_run else ''}{slug} -> {output}")
        update_sitemap(
            repo / "sitemap.xml",
            f"{SITE_ORIGIN}/textile/products/{slug}/",
            date.today().isoformat(),
            args.dry_run,
        )

    if errors:
        print("\nValidation errors:", file=sys.stderr)
        for message in errors:
            print(f"- {message}", file=sys.stderr)
        return 2
    for directory_name, catalog in PRODUCT_CATALOGS.items():
        update_product_catalog(
            repo=repo,
            basics=basics,
            contents=contents,
            image_maps=image_maps,
            directory_name=directory_name,
            catalog_relative_path=catalog["path"],
            default_collection=str(catalog["collection"]),
            dry_run=args.dry_run,
        )
    if not selected:
        print("No products marked YES. Nothing was generated.")
        return 0
    print(f"\nValidated {len(selected)} product(s): {', '.join(selected)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
